"""
Scope PDF → XLSX Converter
Streamlit web app: upload an Xactimate-style depreciation PDF,
download a Garcia-format XLSX ready for import into your tool.
"""

import io
import re
import tempfile
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Scope PDF → XLSX",
    page_icon="📄",
    layout="centered",
)

# ── Styles ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  .title  { font-size: 2rem; font-weight: 700; margin-bottom: 0; }
  .sub    { color: #666; margin-top: 0; margin-bottom: 1.5rem; }
  .room   { background: #f0f4ff; border-left: 4px solid #4a6cf7;
            padding: 0.4rem 0.8rem; border-radius: 4px; margin: 0.5rem 0; font-weight: 600; }
  .badge  { display: inline-block; background: #e8f5e9; color: #2e7d32;
            padding: 2px 10px; border-radius: 12px; font-size: 0.85rem; margin-left: 8px; }
</style>
""", unsafe_allow_html=True)

# ── PDF parsing ───────────────────────────────────────────────────────────────

COL_RANGES = {
    "qty":        (195, 250),
    "unit_price": (270, 315),
    "tax":        (325, 365),
    "o_and_p":    (375, 420),
    "rcv":        (425, 470),
    "deprec":     (490, 530),
    "acv":        (535, 585),
}

ITEM_NUM_RE = re.compile(r"^(\d+)\.$")
QTY_RE      = re.compile(
    r"^([\d,]+\.?\d*)(SF|LF|SY|EA|HR|DA|LS|SQ|MO|WK|GL|LB|CF|CY|TN|PR|BX|RL|BD|CS|PC|FT|IN|MM|CM|M2|M3)$",
    re.I,
)

SKIP_WORDS = {
    "DESCRIPTION", "Totals:", "Total:", "WALLS/CEILING", "MISC", "FLOORING",
    "PLUMBING", "ELECTRICAL", "Grand", "Summary", "Line", "HVAC", "PAINTING",
    "INSULATION", "DRYWALL", "DOORS", "WINDOWS", "CABINETRY", "FINISH",
    "STRUCTURAL", "ROOFING", "CLEANING", "CONTENT",
}


def classify_word(w):
    mid = (w["x0"] + w["x1"]) / 2
    for col, (lo, hi) in COL_RANGES.items():
        if lo <= mid <= hi:
            return col
    return "desc"


def room_name_to_code(name):
    code = name.upper()
    code = re.sub(r"[^A-Z0-9]+", "_", code)
    return code.strip("_")


def infer_activity(desc):
    d = desc.strip()
    if re.match(r"R&R\b", d, re.I) or re.match(r"Remove\s+and\s+Replace", d, re.I):
        return "Remove and Replace"
    if re.match(r"Remove\b", d, re.I):
        return "Remove"
    return "Replace"


def _absorb_row(row_words, item):
    desc_parts = []
    for w in row_words:
        col = classify_word(w)
        txt = w["text"]
        if col in ("desc", "qty"):
            m = QTY_RE.match(txt)
            if m:
                item["qty"]  = float(m.group(1).replace(",", ""))
                item["unit"] = m.group(2).upper()
            else:
                desc_parts.append(txt)
        elif col == "unit_price":
            try:    item["unit_cost"] = float(txt.replace(",", ""))
            except: desc_parts.append(txt)
        elif col == "tax":
            try:    item["sales_tax"] = float(txt.replace(",", ""))
            except: pass
        elif col == "rcv":
            try:    item["rcv"] = float(txt.replace(",", "").strip("()"))
            except: pass
        elif col == "acv":
            try:    item["acv"] = float(txt.replace(",", "").strip("()"))
            except: pass
    if desc_parts:
        item["desc"] = (item["desc"] + " " + " ".join(desc_parts)).strip()


def _parse_page(words, items):
    # Group into rows
    rows = []
    for w in words:
        placed = False
        for row in rows:
            if abs(w["top"] - row["top"]) <= 3:
                row["words"].append(w)
                placed = True
                break
        if not placed:
            rows.append({"top": w["top"], "words": [w]})
    rows.sort(key=lambda r: r["top"])

    # Pre-scan for "RoomName Height:" markers
    room_top_map = {}
    for row in rows:
        texts = [w["text"] for w in row["words"]]
        if "Height:" in texts:
            idx = texts.index("Height:")
            parts = []
            j = idx - 1
            while j >= 0 and not re.match(r"^\d+['\"]", texts[j]):
                parts.insert(0, texts[j])
                j -= 1
            if parts:
                room_top_map[row["top"]] = " ".join(parts)

    current_room = None
    pending = None

    for row in rows:
        top   = row["top"]
        ws    = row["words"]
        texts = [w["text"] for w in ws]
        joined = " ".join(texts)

        if top in room_top_map:
            if pending:
                items.append(pending); pending = None
            current_room = room_top_map[top]
            continue

        if (len(ws) == 1 and texts[0] == "General" and ws[0]["x0"] < 200):
            if pending:
                items.append(pending); pending = None
            current_room = "General"
            continue

        if not current_room:
            continue
        if any(t in SKIP_WORDS for t in texts):
            if pending:
                items.append(pending); pending = None
            continue

        starts = (ws and ITEM_NUM_RE.match(ws[0]["text"]) and ws[0]["x0"] < 50)
        if starts:
            if pending:
                items.append(pending)
            item_num = int(ws[0]["text"].rstrip("."))
            pending = {
                "room_name": current_room,
                "room_code": room_name_to_code(current_room),
                "item_num":  item_num,
                "desc": "", "qty": None, "unit": None,
                "unit_cost": None, "sales_tax": None, "rcv": None, "acv": None,
            }
            _absorb_row(ws[1:], pending)
        elif pending is not None:
            _absorb_row(ws, pending)

    if pending:
        items.append(pending)


def parse_pdf(pdf_bytes):
    items = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if words:
                _parse_page(words, items)
    return items


# ── XLSX builder ─────────────────────────────────────────────────────────────

HEADERS = [
    "#", "Group Code", "Group Description", "Desc", "Age", "Condition",
    "Qty", "Item Amount", "Reported Cost", "Unit Cost", "Unit", "Coverage",
    "Activity", "Worker's Wage", "Labor burden", "Labor Overhead",
    "Material", "Equipment", "Market Conditions", "Labor Minimum",
    "Sales Tax", "RCV", "Life", "Depreciation Type", "Depreciation Amount",
    "Recoverable", "ACV", "Tax", "Replace", "Cat", "Sel", "Owner",
    "Original Vendor", "Date", "Note 1", "Source Name",
]

HDR_FILL = PatternFill("solid", start_color="CCCCCC")
HDR_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_xlsx(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scope"

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

    today = datetime.today()

    for ri, item in enumerate(items, 2):
        qty  = item.get("qty") or 0
        cost = item.get("unit_cost") or 0
        vals = {
            "#": item["item_num"],
            "Group Code": item["room_code"],
            "Group Description": item["room_name"],
            "Desc": item["desc"],
            "Age": 0, "Condition": "Average",
            "Qty": qty,
            "Item Amount": round(qty * cost, 2),
            "Reported Cost": 0,
            "Unit Cost": cost,
            "Unit": item.get("unit") or "",
            "Coverage": "Dwelling",
            "Activity": infer_activity(item["desc"]),
            "Worker's Wage": "", "Labor burden": "", "Labor Overhead": "",
            "Material": "", "Equipment": "", "Market Conditions": "",
            "Labor Minimum": "",
            "Sales Tax": item.get("sales_tax") or 0,
            "RCV": item.get("rcv") or 0,
            "Life": "", "Depreciation Type": "Percent",
            "Depreciation Amount": 0, "Recoverable": "Yes",
            "ACV": item.get("acv") or 0,
            "Tax": "Yes", "Replace": "No",
            "Cat": "", "Sel": "", "Owner": "", "Original Vendor": "",
            "Date": today, "Note 1": "", "Source Name": "",
        }
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(row=ri, column=ci, value=vals[h])
            c.font = BODY_FONT
            c.border = THIN_BORDER
            if h == "Date":
                c.number_format = "MM/DD/YYYY"

    col_widths = {1:5, 2:18, 3:20, 4:45, 5:6, 6:10, 7:8, 8:12,
                  9:12, 10:10, 11:6, 12:10, 13:20}
    for cn, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(cn)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown('<p class="title">📄 Scope PDF → XLSX</p>', unsafe_allow_html=True)
st.markdown('<p class="sub">Upload an Xactimate depreciation scope PDF — get a formatted XLSX back instantly.</p>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Drop your PDF here",
    type=["pdf"],
    label_visibility="collapsed",
)

if uploaded:
    pdf_bytes = uploaded.read()
    filename  = uploaded.name.replace(".pdf", "")

    with st.spinner("Reading PDF and extracting line items…"):
        try:
            items = parse_pdf(pdf_bytes)
        except Exception as e:
            st.error(f"Could not parse PDF: {e}")
            st.stop()

    if not items:
        st.warning("No line items found. Make sure this is an Xactimate-style scope PDF.")
        st.stop()

    # Summary
    rooms = {}
    for it in items:
        rooms.setdefault(it["room_name"], []).append(it)

    st.success(f"✅ Found **{len(items)} line items** across **{len(rooms)} room(s)**")

    # Room breakdown preview
    with st.expander("Preview by room", expanded=True):
        for room, room_items in rooms.items():
            st.markdown(f'<div class="room">{room} <span class="badge">{len(room_items)} items</span></div>', unsafe_allow_html=True)
            preview_data = [
                {
                    "#": it["item_num"],
                    "Description": it["desc"],
                    "Qty": f'{it["qty"]} {it["unit"]}' if it["qty"] else "",
                    "Unit Cost": f'${it["unit_cost"]:,.2f}' if it["unit_cost"] else "",
                    "RCV": f'${it["rcv"]:,.2f}' if it["rcv"] else "",
                    "Activity": infer_activity(it["desc"]),
                }
                for it in room_items
            ]
            st.dataframe(preview_data, hide_index=True, use_container_width=True)

    # Build XLSX
    with st.spinner("Building XLSX…"):
        xlsx_bytes = build_xlsx(items)

    st.download_button(
        label="⬇️  Download XLSX",
        data=xlsx_bytes,
        file_name=f"{filename}_scope.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

else:
    st.info("👆 Upload a PDF above to get started.")
    st.markdown("""
**What this tool does:**
- Reads any Xactimate-style depreciation scope PDF
- Extracts all line items, separated by room
- Outputs a ready-to-import XLSX with the correct 36-column format
- Works even when rooms span multiple pages or appear in different orders
""")
